# -*- coding: utf-8 -*-
"""
Created on Wed Aug 23 11:21:00 2017

@author: Markus
"""


from HIM.utils import *
from HIM import hscClasses
from HIM import plotFunctions as pFun
from HIM import dataHandling as sFun
from openpyxl import load_workbook

#%%

class HSC(object):
    '''
    This class contains the automatic calculation of a predefined supply chain
    After initialization, there are 2 options:
        .calcHSC
        .saveHSC
    '''
    def __init__(self,
                 listHSC,
                 dfTable,
                 listCapacities,
                 totalH2Demand,
                 truckDistance,
                 pipelineDistance,
                 targetCars=100000,
                 beeline=[True, True],
                 industryDemand=0,
                 transportWeight="weightedDistance",
                 learning=True):
        '''
        initialize the hydrogen supply chain:
            Name everything
            prepare the storage partitions
        inputs:
            listHSC: list of hydrogen supply technologies = pathway
            dfTable: dataframe with techno-economic data
            listCapacities: list of hydrogen demand (kg/day) per technology
            totalH2Demand: Total hydrogen demand per year (kg)
            truckDistance: Distance for trucks to drive
            totalLineCost: Pipeline costs
            targetCars: 
            beeline: Boolean of beeline or truck routing is used
            end: Boolean series if the edge is end of the street
            time: Time series of how long the trip for this street is taking
        '''


        #####################################################################
        #Step 1: Initialize the supply chain for the given technologies
        self.dfTable=dfTable
        self.technologies=self.getAttrDict()        
        self.learning=learning
        self.dfHSC=pd.DataFrame(index=listHSC)
        self.dfHSC["kind"]=[self.technologies[key] for key in self.dfHSC.index]
        self.dfHSC["form"]=[dfTable[name].loc["form", key] for (key, name) in zip(self.dfHSC.index, self.dfHSC["kind"])]
        self.dfHSC["capacity"]=listCapacities
        self.getNames()
        self.dfHSC["technology"]=self.dfHSC.index
        self.dfHSC.index=self.dfHSC["name"].values          
        self.dfHSC["intIndex"]=range(len(self.dfHSC.index))
        self.dfHSC["storePart"]=False
        self.dfHSC["TOTEX"]=0
        self.dfHSC["invest"]=0
        self.dfHSC["varOPEX"]=0
        self.dfHSC["fixOPEX"]=0
        self.dfHSC["nTrucks"]=0
        self.dfHSC["pipeLength"]=0
        self.dfHSC["cumCost"]=0
        self.hscClasses={}
        self.industryDemand=industryDemand
        self.totalH2Demand=[totalH2Demand+industryDemand, totalH2Demand]
        self.truckGPD=truckDistance.copy()
        self.pipelineGPD=pipelineDistance.copy()
        self.preStorage=False
        self.targetCars=targetCars
        self.geoms={}
        self.capacityFS=self.dfTable["General"].loc["targetStationSize","General"]
        self.energyDemand=pd.DataFrame(index=["Loss",
                                              "electricityRES",
                                              "electricityGrid",
                                              "NaturalGas",
                                              "Diesel"])
        self.beeline=beeline
        self.transportType=[]
        self.transportWeight=transportWeight
        self.name=''.join((str(e).partition("-")[0][:4] + str(e).partition("-")[2][:2]) for e in self.dfHSC.loc[["Con" not in x for x in self.dfHSC.index], "technology"].values)
        #######################################################################
        #step 2: add important information
        if any(self.dfHSC["kind"]=="Storage"):
            self.preStorage=True
            self.storageIndex=self.dfHSC.loc[self.dfHSC["kind"]=="Storage","intIndex"]
            self.dfHSC.loc[self.dfHSC["intIndex"]==self.storageIndex[0]-1,"storePart"]=True
            self.dfHSC.loc[self.dfHSC["intIndex"]==self.storageIndex[0]+1,"storePart"]=True
    
    #get dictionary of attributes from the ImportTablesTechnology.xslx
    def getAttrDict(self):
        attr={}
        for key in self.dfTable:
            if "lr6" in key:
                continue
            for tech in self.dfTable[key].columns:
                attr[tech]=key

        return attr
    
    def getNames(self):
        names=[]
        for (key, name) in zip(self.dfHSC.index, self.dfHSC["kind"]):
            count=len(self.dfHSC[self.dfHSC["kind"]==name]["kind"])
            if count<=1:
                names.append(name)
            else:
                x=1
                while name + str(x) in names:
                    x+=1
                names.append(name+str(x))
        self.dfHSC["name"]=names    
    
    def calcFSCost(self ,targetNumber, learning = False):
        '''
        calculate Fuelstation cost based on learning rate and scaleup
        '''
        baseCost=self.dfTable["General"].loc["baseStationCost","General"]
        baseSize=self.dfTable["General"].loc["baseStationSize","General"]
        baseNumber=self.dfTable["General"].loc["baseStationNumber","General"]
        learningRate=self.dfTable["General"].loc["learningRate","General"]
        
        if learning:
            V0=baseSize*baseNumber
            V1=self.capacityFS*targetNumber
            beta=np.log2(1-learningRate)
            learningFactor=((V1/V0)**beta)/(1+beta)
        else:
            learningFactor=1
        
        self.dfTable["Station"].loc["stationInvest",:]=baseCost*self.dfTable["Station"].loc["stationMult",:]*(self.capacityFS/baseSize)**self.dfTable["Station"].loc["stationScaling",:]*learningFactor
                      
    def calcHSC(self,
                cumCost=0,
                abstract=False):
        '''
        calculates the hydrogen supply chain:
            1. initialize the chain parts
            2. calculate the results
            3. extract the results
            Resulting dataframe: HSC.dfHSCRes
        '''
        self.startCost=cumCost
        iTrans=0
        for (key, (kind, technology, storagePartition, capacity, index)) in self.dfHSC.loc[:,[ "kind","technology","storePart", "capacity", "intIndex"]].iterrows():
            if kind=="Production":
                self.hscClasses[key]=hscClasses.Production(capacity,
                                                            technology,
                                                            self.dfTable)
            elif kind=="Import":
                self.hscClasses[key]=hscClasses.Import(capacity,
                                                       technology,
                                                       self.dfTable)
            elif kind=="Storage":
                self.hscClasses[key]=hscClasses.Storage(capacity,
                                                         technology,
                                                         self.dfTable,
                                                         costH2In=cumCost)
                self.preStorage=False
            elif kind=="Transport":
                #check if pipeline or Truck
                if technology=="Pipeline":
                    totalLineCost=self.pipelineGPD[iTrans]["lineCost"].sum()*1e6
                    self.dfHSC.loc[key,"pipeLength"]=self.pipelineGPD[iTrans][self.transportWeight].sum()
                    #self.geoms[key]=self.pipelineDistance[iTrans].geometry
                    self.hscClasses[key]=hscClasses.Pipeline(self.totalH2Demand[iTrans],
                                                              totalLineCost,
                                                              self.pipelineGPD[iTrans]["weightedDistance"].sum()*1e3,
                                                              self.dfTable,
                                                              costH2In=cumCost)
                    self.transportType.append("Pipeline")
                else:
                    #self.geoms[key]=self.truckDistance[iTrans].geometry
                    self.hscClasses[key]=hscClasses.Truck(capacity,
                                                           self.truckGPD[iTrans][self.transportWeight],
                                                           technology,
                                                           self.dfTable,
                                                           costH2In=cumCost,
                                                           beeline=self.beeline[iTrans],
                                                           end=self.truckGPD[iTrans]["edge"],
                                                           time=self.truckGPD[iTrans]["time"],
                                                           totalH2Demand=self.totalH2Demand[iTrans])
                    self.truckGPD[iTrans]["TOTEX"]=self.hscClasses[key].getTOTEX()
                    self.truckGPD[iTrans]["nTruckPerDay"]=self.hscClasses[key].nTruckPerDay
                    
                    self.transportType.append("Truck")
                    
                iTrans+=1
                
            elif kind=="Station":
                if "learningRate" in self.dfTable["General"].index:
                    self.calcFSCost(len(capacity), learning = self.learning)
                self.hscClasses[key]=hscClasses.Station(capacity,
                                                         technology,
                                                         self.dfTable,
                                                         costH2In=cumCost)
            

            
            
            elif kind=="Connector":
            #The conversion technology needs additional information about the
            #inlet and outlet pressure, in case it is an compressor
                pressureIn=0
                pressureOut=0
                nextTech=0                
                if technology=="Compressor":
                    #previous Technology --> elaborating starting pressure
                    prevKind=self.dfHSC.loc[[x == index-1 for x in self.dfHSC["intIndex"]],"kind"][0]
                    prevTech=self.dfHSC.loc[[x == index-1 for x in self.dfHSC["intIndex"]],"technology"][0]
                    i=1
                    while prevTech=="None" and index-i>0:
                        i+=1
                        prevKind=self.dfHSC.loc[[x == index-i for x in self.dfHSC["intIndex"]],"kind"][0]
                        prevTech=self.dfHSC.loc[[x == index-i for x in self.dfHSC["intIndex"]],"technology"][0]
                    #since conversion pressure are not fixed: try method
                    try:
                        pressureIn=self.dfTable[prevKind].loc["pressureOut",prevTech]
                    except:
                        if prevTech=="Dehydrogenation":
                            pressureIn=2.
                        else: pressureIn=999.
                    #following Technology --> elaborating starting pressure
                    nextKind=self.dfHSC.loc[[x == index+1 for x in self.dfHSC["intIndex"]],"kind"][0]
                    nextTech=self.dfHSC.loc[[x == index+1 for x in self.dfHSC["intIndex"]],"technology"][0]
                    i=1
                    while nextTech=="None":
                        i+=1
                        nextKind=self.dfHSC.loc[[x == index+i for x in self.dfHSC["intIndex"]],"kind"][0]
                        nextTech=self.dfHSC.loc[[x == index+i for x in self.dfHSC["intIndex"]],"technology"][0]                  
                    
                    try: 
                        pressureOut=self.dfTable[nextKind].loc["pressureIn",nextTech]
                    except:
                        #Input pressure for Liquefaction as well as Hydrogenation as set to 30 bar
                        pressureOut=30.
                    #if pressure in > pressure Out --> No compression necessary
                    pressureOut=max(pressureIn, pressureOut)
                    
                
                
                if self.preStorage:
                    self.hscClasses[key]=hscClasses.Connector(capacity,
                                                               technology, 
                                                               self.dfTable,
                                                               costH2In=cumCost,
                                                               pressureIn=pressureIn,
                                                               pressureOut=pressureOut,
                                                               nextStep=nextTech,
                                                               storagePartition=storagePartition)
                else:
                    self.hscClasses[key]=hscClasses.Connector2(capacity,
                                                                technology,
                                                                self.dfTable,
                                                                costH2In=cumCost,
                                                                pressureIn=pressureIn,
                                                                pressureOut=pressureOut,
                                                                nextStep=nextTech,
                                                                storagePartition=storagePartition)
                
               
                
                
            cumCost+=self.hscClasses[key].getMeanTOTEX()
            self.dfHSC.loc[key, "TOTEX"]=self.hscClasses[key].getMeanTOTEX()
            self.dfHSC.loc[key, "invest"]=self.hscClasses[key].getTotalInvest()

            self.dfHSC.loc[key, "CAPEX"]=self.hscClasses[key].getMeanValue(self.hscClasses[key].CAPEX)
            self.dfHSC.loc[key, "fixOPEX"]=self.hscClasses[key].getMeanValue(self.hscClasses[key].fixOPEX)
            self.dfHSC.loc[key, "varOPEX"]=self.hscClasses[key].getMeanValue(self.hscClasses[key].varOPEX)
            self.dfHSC.loc[key, "nTrucks"]=self.hscClasses[key].numberOfTrucks
            self.dfHSC.loc[key, "cumCost"]=cumCost
            self.dfHSC.loc[key, "additionalEmissions"]=self.hscClasses[key].CO2Emissions
            self.dfHSC.loc[key, "additionalPE"]=self.hscClasses[key].primary
            self.energyDemand[key]=self.hscClasses[key].getDemand()
        #Add the energy demand to the results 
        self.dfHSC=pd.concat([self.dfHSC, self.energyDemand.T],axis=1)

        ##Test!!!
        self.dfHSC.loc[self.dfHSC.index[0],"TOTEX"]=self.dfHSC.loc[self.dfHSC.index[0],"TOTEX"]+self.startCost
        
        #Exclude everything in the same way as for H2Mobility
        self.dfHSCRes=self.dfHSC.loc[:,["TOTEX",
                                        "CAPEX",
                                        "fixOPEX",
                                        "varOPEX",
                                        "invest",
                                        "cumCost",
                                        "technology",
                                        "nTrucks",
                                        "pipeLength",
                                        "Loss",
                                        "electricityRES",
                                        "electricityGrid",
                                        "NaturalGas",
                                        "Diesel"]]
        self.calcLossesRecursive()
        self.calcEmissions()
        self.calcPrimaryDemand()        
        self.dfHSCRes=self.dfHSCRes.round(4)
#%%
    def createResFolders(self, pathResults, savePlot=False, saveResults=False):
        self.savePlot=savePlot
        self.saveResults=saveResults
        self.pathResults=os.path.join(pathResults, self.name)
        if savePlot:
            self.pathPlot=os.path.join(self.pathResults, "Graphics")
            os.makedirs(self.pathPlot)
        if saveResults:
            self.pathRes=os.path.join(self.pathResults, "data")
            os.makedirs(self.pathRes)
    
    def calcEmissions(self, useOverCapacity=True):
        '''
        calculate emissions based on input Values in dfTable
        '''
        if useOverCapacity:
            factor=self.dfHSCRes["overCapacity"]
        else:
            factor=1
        emissions=self.dfTable["General"].loc[["emissionDiesel",
                                              "emissionNG",
                                              "emissionGrid",
                                              "emissionRES"], "General"]
        emissionDiesel=emissions["emissionDiesel"]*self.dfHSC["Diesel"]
        emissionNG=emissions["emissionNG"]*self.dfHSC["NaturalGas"]
        emissionGrid=emissions["emissionGrid"]*self.dfHSC["electricityGrid"]
        emissionRES=emissions["emissionRES"]*self.dfHSC["electricityRES"]
        self.dfHSCRes["CO2Emissions[kg/kg]"]=factor*((emissionDiesel+emissionNG+emissionGrid+emissionRES)/1000+self.dfHSC["additionalEmissions"])

    def calcPrimaryDemand(self, useOverCapacity=True):
        '''
        calculate emissions based on input Values in dfTable
        '''
        if useOverCapacity:
            factor=self.dfHSCRes["overCapacity"]
        else:
            factor=1
        primaryDemand=self.dfTable["General"].loc[["primaryDiesel",
                                              "primaryNG",
                                              "primaryGrid",
                                              "primaryRES"], "General"]
        primaryDiesel=primaryDemand["primaryDiesel"]*self.dfHSC["Diesel"]
        primaryNG=primaryDemand["primaryNG"]*self.dfHSC["NaturalGas"]
        primaryGrid=primaryDemand["primaryGrid"]*self.dfHSC["electricityGrid"]
        primaryRES=primaryDemand["primaryRES"]*self.dfHSC["electricityRES"]
        self.dfHSCRes["primaryEnergy[MJ/kg]"]=factor*(primaryDiesel+primaryNG+primaryGrid+primaryRES+self.dfHSC["additionalPE"])
    #%%
    def calcLossesRecursive(self):
        '''
        recursive calculate losses back to get an overcapacity that is necessary
        '''
        loss=[1]
        lossesSingle=self.dfHSCRes["Loss"]
        for i in range(lossesSingle.size-1):
            loss.append(lossesSingle[-1*(i+1)]*loss[i])
        loss.reverse()
        self.dfHSCRes["overCapacity"]=loss
#%%
        '''
        saves the Hydrogen supply chain results to the given excel-file
        '''
    def saveHSC(self, pathData, i, name="HSCRes.xlsx"):
        self.filePath=os.path.join(pathData,name)
        self.sheetName=str(i)
        if path.isfile(self.filePath):
            book = load_workbook(self.filePath)
            if self.sheetName in book.sheetnames:
                std=book.get_sheet_by_name(self.sheetName)
                book.remove_sheet(std)
            writer=pd.ExcelWriter(self.filePath, engine = 'openpyxl')
            writer.book=book
            self.dfHSCRes.to_excel(writer,
                                   sheet_name=self.sheetName)
            writer.save()
            writer.close()  
        else:
            self.dfHSCRes.to_excel(self.filePath, sheet_name=self.sheetName)

  
#%%    
    def plotHSC(self,
                figsize=(14,10),
                background=None,
                source=None,
                hub=None,
                sink=None,
                truckVar="capacity",
                pipeVar="capacity",
                zorder=3,
                alpha=1,
                savePlot=False,
                pathPlot=None,
                show=True):
        '''
        plot the scenario as a map of the investigated country
        '''
        bg_area=(87/255, 133/255, 147/255)
        bg_lines=(99/255, 150/255, 167/255)
        
        self.fig,self.ax=plt.subplots(figsize=figsize)
        self.ax.set_aspect('equal')
        self.ax.axis("off")
        rangeTransport=len(self.transportType)
        if isinstance(background, gpd.GeoDataFrame):
            pFun.plot_polygon_collection(ax=self.ax,
                                 geoms=background.geometry,
                                 colors_or_values=[bg_area for ix in background.index],
                                 plot_values=False,
                                 vmin=None,
                                 vmax=None,
                                 cmap=None,
                                 edgecolor=bg_lines,
                                 alpha=1,
                                 label="Administrative Area")
        #Sinks
        if isinstance(sink, gpd.GeoDataFrame):
            pFun.plotGPDPoints(sink, self.ax,
                               color="black",
                               label="Fueling Stations")
        
        #Source
        sourceMarkerMax=100
        sourceMarkerMin=1
        if isinstance(source, gpd.GeoDataFrame):
            if self.transportType[0]=="Pipeline":
                source["markersize"]=((source["pipe_kt_a"]-source["pipe_kt_a"].min())/source["pipe_kt_a"].max()*(sourceMarkerMax-sourceMarkerMin)+sourceMarkerMin)
            else:
                source["markersize"]=((source["truck_kt_a"]-source["truck_kt_a"].min())/source["truck_kt_a"].max()*(sourceMarkerMax-sourceMarkerMin)+sourceMarkerMin)
            pFun.plotGPDPoints(source,self.ax,
                               colName="markersize",
                               color=(178/255, 223/255, 138/255),
                               zorder=zorder+1,
                               marker="D",
                               label="Electrolyzer\n %.1f" % source["pipe_kt_a"].min() + " - %.1f" % source["pipe_kt_a"].max() + "kt/a")

        if self.industryDemand>0:
            indMarkerSizeMax=100
            indMarkerSizeMin=0
            if isinstance(hub, gpd.GeoDataFrame):
                hub["markersize"]=((hub["H2Ind_kt"]-hub["H2Ind_kt"].min())/hub["H2Ind_kt"].max()*(indMarkerSizeMax-indMarkerSizeMin)+indMarkerSizeMin)
                pFun.plotGPDPoints(hub,self.ax,
                                   colName="markersize",
                                   color="blue",
                                   zorder=zorder+1,
                                   marker="^",
                                   label="Industrial Demand\n %.1f" % hub["H2Ind_kt"].min() + " - %.1f" % hub["H2Ind_kt"].max() + "kt/a")        

        #Transport
        lineWidthMax=[5,0.5]
        lineWidthMin=[0.5,0.5]
        colorStyle=["redToWhite", "black"]
        maxRange=[5,1]
        pipeLabel=["Transmission Pipeline", "Distribution Pipeline"]
        truckLabel=["Truck Routes", "Truck Routes (Distribution)"]
        for i in range(rangeTransport):
            if self.transportType[i]=="Pipeline":
                pFun.plotGpdLinesVar(self.pipelineGPD[i],
                                     pipeVar,
                                     self.ax,
                                     zorder=zorder-i,
                                     alpha=alpha,
                                     name=colorStyle[i],
                                     rangeMax=maxRange[i],
                                     maxLineWidth=lineWidthMax[i],
                                     minLineWidth=lineWidthMin[i],
                                     label=pipeLabel[i])
            else:
                pFun.plotGpdLinesVar(self.truckGPD[i],
                                     truckVar,
                                     self.ax,
                                     zorder=zorder-i,
                                     alpha=alpha,
                                     name="black",
                                     rangeMax=maxRange[i],
                                     maxLineWidth=lineWidthMax[i],
                                     minLineWidth=lineWidthMin[i],
                                     label=truckLabel[i])
                

        plt.legend(bbox_to_anchor=(1.05, 1), loc=2, borderaxespad=0.)           
        if savePlot:
            plt.savefig(os.path.join(pathPlot,self.name), bbox_inches="tight")
        if show:
            plt.show()
        
